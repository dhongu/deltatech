# © 2026 Deltatech
# See README.rst file on addons root folder for license details

import base64
import csv
from io import BytesIO, StringIO

from odoo import _, fields, models
from odoo.exceptions import UserError


class DeliveryVendorDocumentImportWizard(models.TransientModel):
    _name = "delivery.vendor.document.import.wizard"
    _description = "Import Delivery Document Lines"

    file = fields.Binary(string="File", required=True, help="Upload an Excel (.xlsx) or CSV file")
    filename = fields.Char(string="Filename")
    file_type = fields.Selection(
        [
            ("xlsx", "Excel (.xlsx)"),
            ("csv", "CSV"),
        ],
        string="File Type",
        default="xlsx",
        required=True,
    )
    product_match_by = fields.Selection(
        [
            ("default_code", "Internal Reference"),
            ("barcode", "Barcode"),
            ("name", "Name"),
        ],
        string="Match Product By",
        default="default_code",
        required=True,
    )
    delimiter = fields.Char(string="CSV Delimiter", help="Used only for CSV; leave empty to auto-detect", size=1)

    def action_import(self):
        self.ensure_one()
        doc = self.env["delivery.vendor.document"].browse(self.env.context.get("active_id"))
        if not doc:
            raise UserError(_("No active delivery document found."))
        if doc.state != "draft":
            raise UserError(_("You can import lines only in Draft state."))

        rows = self._read_rows()
        if not rows:
            raise UserError(_("The file does not contain any rows."))

        # expected columns: product, quantity, (optional) uom, (optional) price_unit, (optional) name
        errors = []
        for index, row in enumerate(rows, start=2):  # assume header at row 1
            try:
                product_key = (row.get("product") or "").strip()
                qty_val = row.get("quantity")
                uom_key = (row.get("uom") or "").strip()
                price_val = row.get("price_unit")
                name = (row.get("name") or "").strip()

                if not product_key:
                    raise UserError(_("Row %s: product is required.") % index)
                if qty_val in (None, ""):
                    raise UserError(_("Row %s: quantity is required.") % index)
                try:
                    qty = float(qty_val)
                except Exception as err:
                    raise UserError(_("Row %s: quantity is not a number.") % index) from err
                if qty <= 0:
                    raise UserError(_("Row %s: quantity must be > 0.") % index)

                product = self._match_product(product_key)
                uom = product.uom_id
                if uom_key:
                    uom = self.env["uom.uom"].search([("name", "=", uom_key)], limit=1) or uom

                price_unit = None
                if price_val not in (None, ""):
                    try:
                        price_unit = float(price_val)
                    except Exception as err:
                        raise UserError(_("Row %s: price_unit is not a number.") % index) from err

                vals = {
                    "document_id": doc.id,
                    "product_id": product.id,
                    "name": name or product.display_name,
                    "quantity": qty,
                    "product_uom": uom.id,
                }
                if price_unit is not None:
                    vals["price_unit"] = price_unit

                self.env["delivery.vendor.document.line"].create(vals)
            except Exception as e:
                errors.append(str(e))

        if errors:
            raise UserError("\n".join(errors))

        return {
            "type": "ir.actions.act_window",
            "res_model": "delivery.vendor.document",
            "res_id": doc.id,
            "view_mode": "form",
            "target": "current",
        }

    def _read_rows(self):
        if self.file_type == "xlsx":
            try:
                import openpyxl  # noqa: F401
            except Exception as err:
                raise UserError(_('XLSX support requires the "openpyxl" Python package.')) from err
            content = base64.b64decode(self.file or b"")
            from openpyxl import load_workbook

            wb = load_workbook(filename=BytesIO(content), read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter)
            headers = [h.strip() if isinstance(h, str) else h for h in headers]
            result = []
            for r in rows_iter:
                row = {}
                for i, h in enumerate(headers):
                    if not h:
                        continue
                    val = r[i] if i < len(r) else None
                    row[h] = (
                        val
                        if val is None or isinstance(val, str)
                        else (str(val) if isinstance(val, int | float) else val)
                    )
                result.append(row)
            return result
        # CSV
        data = base64.b64decode(self.file or b"")
        text = data.decode("utf-8")
        if self.delimiter:
            reader = csv.DictReader(StringIO(text), delimiter=self.delimiter)
        else:
            try:
                dialect = csv.Sniffer().sniff(text.splitlines()[0])
                reader = csv.DictReader(StringIO(text), dialect=dialect)
            except Exception:
                reader = csv.DictReader(StringIO(text))
        return list(reader)

    def _match_product(self, key):
        key = str(key).strip()
        Product = self.env["product.product"]
        if self.product_match_by == "default_code":
            product = Product.search([("default_code", "=", key)], limit=1)
        elif self.product_match_by == "barcode":
            product = Product.search([("barcode", "=", key)], limit=1)
        else:
            product = Product.search([("name", "=", key)], limit=1)
        if not product:
            raise UserError(_('Product "%s" not found.') % key)
        return product
